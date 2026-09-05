import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.database import get_districts, get_villages

def generate_resurvey_excel() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resurvey Progress Summary"
    
    # Styles
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill_1 = PatternFill(start_color="205493", end_color="205493", fill_type="solid")
    header_fill_2 = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_fill_3 = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")
    header_fill_4 = PatternFill(start_color="C25E00", end_color="C25E00", fill_type="solid")
    
    data_font = Font(name="Calibri", size=10)
    total_font = Font(name="Calibri", size=10, bold=True)
    total_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Row 1: Title
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "100% Completed Villages of Non Cadastral & Cadastral Category as on 04.09.2026"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row 2 & 3: Multi-level Headers
    ws.merge_cells("A2:A3")
    ws["A2"] = "SI.No"
    ws.merge_cells("B2:B3")
    ws["B2"] = "District"
    
    ws.merge_cells("C2:D2")
    ws["C2"] = "No of Resurvey villages"
    ws["C3"] = "Non-Cadastral villages (373)"
    ws["D3"] = "Cadastral villages (70)"
    
    ws.merge_cells("E2:G2")
    ws["E2"] = "GT Completed Villages of Non Cadastral"
    ws["E3"] = "No. of Non-Cadastral"
    ws["F3"] = "Name of the Village & Mandal"
    ws["G3"] = "TOTAL Extent"
    
    ws.merge_cells("H2:J2")
    ws["H2"] = "GT Completed Villages of Cadastral"
    ws["H3"] = "No. of Cadastral"
    ws["I3"] = "Name of the Village & Mandal"
    ws["J3"] = "TOTAL Extent"
    
    ws.merge_cells("K2:K3")
    ws["K2"] = "Shape files sent to CSO"
    ws.merge_cells("L2:L3")
    ws["L2"] = "Remarks"

    for row in range(2, 4):
        ws.row_dimensions[row].height = 24
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col in [1, 2]:
                cell.fill = header_fill_1
            elif col in [3, 4]:
                cell.fill = header_fill_1
            elif col in [5, 6, 7]:
                cell.fill = header_fill_2
            elif col in [8, 9, 10]:
                cell.fill = header_fill_3
            else:
                cell.fill = header_fill_4

    # Data Population
    districts = get_districts()
    all_villages = get_villages()
    
    current_row = 4
    for idx, d in enumerate(districts, start=1):
        d_name = d["name"]
        d_vlgs = [v for v in all_villages if v.get("district_name", "").lower() == d_name.lower()]
        non_cad_vlgs = [v for v in d_vlgs if v.get("category") == "Non-Cadastral"]
        cad_vlgs = [v for v in d_vlgs if v.get("category") == "Cadastral"]
        
        max_subrows = max(len(non_cad_vlgs), len(cad_vlgs), 1)
        
        for sub_idx in range(max_subrows):
            nc_v = non_cad_vlgs[sub_idx] if sub_idx < len(non_cad_vlgs) else None
            c_v = cad_vlgs[sub_idx] if sub_idx < len(cad_vlgs) else None
            
            # Col A: SI.No (only on first subrow)
            if sub_idx == 0:
                ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=2, value=d_name)
                ws.cell(row=current_row, column=3, value=d.get("non_cadastral_target", 0)).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=4, value=d.get("cadastral_target", 0)).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=5, value=len(non_cad_vlgs)).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=8, value=len(cad_vlgs)).alignment = Alignment(horizontal="center")
            else:
                ws.cell(row=current_row, column=1, value="")
                ws.cell(row=current_row, column=2, value=d_name)
                ws.cell(row=current_row, column=3, value="")
                ws.cell(row=current_row, column=4, value="")
                ws.cell(row=current_row, column=5, value="")
                ws.cell(row=current_row, column=8, value="")
            
            # Non-cadastral village name & extent
            if nc_v:
                v_str = f"{nc_v.get('village_name', '')}, {nc_v.get('mandal_name', '')}".strip(', ')
                ws.cell(row=current_row, column=6, value=v_str)
                ws.cell(row=current_row, column=7, value=nc_v.get("extent_raw", "")).alignment = Alignment(horizontal="right")
                
                # Shapefile & remarks if nc_v has them
                sf_text = f"Shape file {nc_v.get('shapefile_status')}" if nc_v.get("shapefile_status") != "Pending" else ""
                ws.cell(row=current_row, column=11, value=sf_text)
                ws.cell(row=current_row, column=12, value=nc_v.get("remarks", ""))
            
            # Cadastral village name & extent
            if c_v:
                v_str = f"{c_v.get('village_name', '')}, {c_v.get('mandal_name', '')}".strip(', ')
                ws.cell(row=current_row, column=9, value=v_str)
                ws.cell(row=current_row, column=10, value=c_v.get("extent_raw", "")).alignment = Alignment(horizontal="right")
                
                # Override shapefile/remarks if present on cadastral row
                if not nc_v:
                    sf_text = f"Shape file {c_v.get('shapefile_status')}" if c_v.get("shapefile_status") != "Pending" else ""
                    ws.cell(row=current_row, column=11, value=sf_text)
                    ws.cell(row=current_row, column=12, value=c_v.get("remarks", ""))

            # Borders & fonts
            for col in range(1, 13):
                c = ws.cell(row=current_row, column=col)
                c.font = data_font
                c.border = thin_border
            
            current_row += 1
            
    # Summary Total Row
    ws.cell(row=current_row, column=1, value="Total")
    ws.cell(row=current_row, column=2, value="All Districts (32)")
    ws.cell(row=current_row, column=3, value=sum(d.get("non_cadastral_target", 0) for d in districts))
    ws.cell(row=current_row, column=4, value=sum(d.get("cadastral_target", 0) for d in districts))
    ws.cell(row=current_row, column=5, value=len([v for v in all_villages if v.get("category") == "Non-Cadastral"]))
    ws.cell(row=current_row, column=8, value=len([v for v in all_villages if v.get("category") == "Cadastral"]))
    ws.cell(row=current_row, column=11, value=len([v for v in all_villages if v.get("sent_to_cso") or v.get("shapefile_status") == "Completed"]))
    ws.cell(row=current_row, column=12, value="Statewide Progress Live Sync")
    
    for col in range(1, 13):
        c = ws.cell(row=current_row, column=col)
        c.font = total_font
        c.fill = total_fill
        c.border = thin_border
        if col in [1, 3, 4, 5, 8, 11]:
            c.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = {1: 8, 2: 20, 3: 16, 4: 16, 5: 14, 6: 34, 7: 15, 8: 14, 9: 34, 10: 15, 11: 22, 12: 45}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
